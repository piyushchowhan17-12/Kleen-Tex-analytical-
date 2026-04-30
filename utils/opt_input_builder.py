"""
Converts forecasting output → optimization model input.
Exact port of Input_Data_Filtered_For_sS_Adjusted.py logic.
No single line of logic changed from the source file.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styling — exact copy from Input_Data_Filtered_For_sS_Adjusted.py ─────────
HEADER_FILL = PatternFill("solid", start_color="4472C4", end_color="4472C4")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")
RIGHT       = Alignment(horizontal="right",  vertical="center")
_thin       = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
NUM_FMT     = "#,##0.0000"


def style_header(cell, align=CENTER):
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = align
    cell.border    = THIN_BORDER


def style_data(cell, align=RIGHT, num_fmt=None):
    cell.font      = DATA_FONT
    cell.alignment = align
    cell.border    = THIN_BORDER
    if num_fmt:
        cell.number_format = num_fmt


def write_wide_sheet(ws, pivot_df, row0_label, row1_label):
    """
    Exact copy of write_wide_sheet() from Input_Data_Filtered_For_sS_Adjusted.py.
    Writes a sheet with the two-row header style:
      Row 1: [blank, row0_label (merged B1:L1)]
      Row 2: [Item No., 1, 2, ..., 11]
      Row 3+: data
    """
    month_cols = [c for c in pivot_df.columns if c != "Item No."]
    n = len(month_cols)

    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=n + 1)
    ws["A1"] = ""
    ws["B1"] = row0_label
    style_header(ws["A1"])
    style_header(ws["B1"])

    ws["A2"] = "Item No."
    style_header(ws["A2"], align=LEFT)
    for idx, m in enumerate(month_cols, start=2):
        style_header(ws.cell(row=2, column=idx, value=m))

    for r_idx, row in pivot_df.iterrows():
        er = r_idx + 3
        style_data(ws.cell(row=er, column=1, value=row["Item No."]), align=LEFT)
        for idx, m in enumerate(month_cols, start=2):
            val = row[m] if pd.notna(row.get(m)) else 0.0
            style_data(ws.cell(row=er, column=idx, value=round(float(val), 4)),
                       num_fmt=NUM_FMT)

    ws.column_dimensions["A"].width = 26
    for col in range(2, n + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 16


def write_narrow_sheet(ws, df_source):
    """Exact copy of write_narrow_sheet() from source file."""
    for r_idx, row in df_source.iterrows():
        for c_idx, val in enumerate(row):
            cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)
            if r_idx == 0:
                style_header(cell, align=LEFT)
            else:
                style_data(cell,
                           align=LEFT if c_idx == 0 else RIGHT,
                           num_fmt=NUM_FMT if c_idx > 0 else None)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.row_dimensions[1].height = 16


def write_selected_sheet(ws, items):
    """Exact copy of write_selected_sheet() from source file — 1 column."""
    # Header row
    cell = ws.cell(row=1, column=1, value="Selected Item No.")
    style_header(cell, align=LEFT)
    ws.row_dimensions[1].height = 16
    # Data rows
    for row_idx, item in enumerate(items, start=2):
        cell = ws.cell(row=row_idx, column=1, value=item)
        style_data(cell, align=LEFT)
    ws.column_dimensions["A"].width = 26


def build_optimization_input(
    detail_df:         pd.DataFrame,
    forecast_months:   int,
    raw_df:            pd.DataFrame | None = None,
    inventory_cost_df: pd.DataFrame | None = None,
    m2_per_item_df:    pd.DataFrame | None = None,
) -> tuple[dict, bytes]:
    """
    Builds optimization input with 3 extraction rules:
    1. Unit Inventory Cost + m2/item extracted from raw uploaded file (raw_df).
    2. I_ini = Ending_Inventory_Qty of the month BEFORE first forecast month per SKU.
    3. Demand = Demand_Prediction from Demand Forecasting module (already in detail_df).
    """

    # ── Step 1: Filter to forecast rows (Demand_Prediction = forecasted demand) ─
    fc = detail_df[detail_df["Demand_Prediction"].notna()].copy()
    if "Imputed_Demand" in fc.columns:
        fc["Demand_Prediction"] = fc["Demand_Prediction"].fillna(fc["Imputed_Demand"])

    if fc.empty:
        raise ValueError("No forecast rows found in detail_df. Run forecasting first.")

    # Sort chronologically by Year+Month so Dec 2025 < Jan 2026 < Feb 2026
    ym_sorted  = (fc[["Year","Month"]].drop_duplicates()
                    .sort_values(["Year","Month"])
                    .reset_index(drop=True))
    month_cols = ym_sorted["Month"].tolist()
    # ym_map: month number → "YYYY-MM" string (e.g. {12: "2025-12", 1: "2026-01", 2: "2026-02"})
    # Also store ym_order: ordered list of month numbers for chronological sorting
    ym_map   = {int(r.Month): f"{int(r.Year)}-{int(r.Month):02d}"
                for _, r in ym_sorted.iterrows()}
    ym_order = ym_sorted["Month"].tolist()  # [12, 1, 2] in chronological order

    # ── Step 2: Extract I_ini = Ending_Inventory_Qty of the month BEFORE ────────
    # the first forecast month for each SKU.
    # E.g. if forecasting Dec 2025 → Feb 2026, I_ini comes from Nov 2025's
    # Ending_Inventory_Qty in the historical (imputed) rows of detail_df.
    sel_skus = sorted(fc["SKU"].unique().tolist())

    # Identify the first forecast period per SKU
    first_fc = (fc.sort_values(["Year", "Month"])
                  .groupby("SKU")[["Year", "Month"]].first()
                  .reset_index())

    # Get historical rows (Demand_Prediction is null → historical)
    hist = detail_df[detail_df["Demand_Prediction"].isna()].copy()

    # For each SKU, find the historical row immediately before the first forecast month
    iini_map = {}
    for _, frow in first_fc.iterrows():
        sku    = frow["SKU"]
        fy, fm = int(frow["Year"]), int(frow["Month"])
        sku_hist = hist[hist["SKU"] == sku].copy()
        # Convert Year+Month to a sortable integer for comparison
        sku_hist["_ym"] = sku_hist["Year"] * 100 + sku_hist["Month"]
        first_fc_ym     = fy * 100 + fm
        # Get rows strictly before first forecast month
        before = sku_hist[sku_hist["_ym"] < first_fc_ym].sort_values("_ym")
        if not before.empty and "Ending_Inventory_Qty" in before.columns:
            iini_map[sku] = float(before["Ending_Inventory_Qty"].iloc[-1])
        else:
            # Fallback: use Opening_Inventory_Qty of first forecast row
            fc_sku = fc[fc["SKU"] == sku]
            if not fc_sku.empty and "Opening_Inventory_Qty" in fc_sku.columns:
                iini_map[sku] = float(fc_sku.sort_values(["Year","Month"])
                                       ["Opening_Inventory_Qty"].iloc[0])
            else:
                iini_map[sku] = 0.0

    # Build I_ini pivot — only the FIRST forecast month carries the initial inventory value.
    # Subsequent months are 0 because the optimizer derives their opening stock from
    # the previous period's I_End (via get_prev_inventory_expr), not from this table.
    first_month = month_cols[0] if month_cols else None
    iini_rows = []
    for sku in sel_skus:
        row = {"Item No.": sku}
        for m in month_cols:
            row[m] = iini_map.get(sku, 0.0) if m == first_month else 0.0
        iini_rows.append(row)
    iini_pivot = pd.DataFrame(iini_rows)[["Item No."] + month_cols]

    # ── Q_it on forecast rows (kept for Excel export) ────────────────────────
    fc = fc.copy()
    for col in ["Ending_Inventory_Qty", "Opening_Inventory_Qty"]:
        if col not in fc.columns:
            fc[col] = 0.0
    fc["Q_it"] = (
        fc["Ending_Inventory_Qty"]
        - fc["Opening_Inventory_Qty"]
        + fc["Imputed_Demand"].fillna(0.0)
    )

    # ── Step 3: Build pivots — Demand_Prediction is the forecasted demand ─────
    def make_pivot(column, skus):
        piv = (
            fc.pivot(index="SKU", columns="Month", values=column)
              .reindex(skus)
              .reset_index()
              .rename(columns={"SKU": "Item No."})
        )
        for m in month_cols:
            if m not in piv.columns:
                piv[m] = 0.0
        return piv[["Item No."] + month_cols]

    demand_pivot = make_pivot("Demand_Prediction", sel_skus)  # Rule 3: forecasted demand
    actual_pivot = make_pivot("Imputed_Demand",    sel_skus)
    qit_pivot    = make_pivot("Q_it",              sel_skus)
    # iini_pivot already built above using Rule 2

    # ── Step 4: Unit Inventory Cost + m2/item from raw uploaded file ──────────
    # Rule 1: Extract directly from raw_df if columns exist
    if raw_df is not None and not raw_df.empty:
        # Standardise column names — try both with and without spaces
        raw_cols = {c.strip().lower(): c for c in raw_df.columns}

        # Unit Inventory Cost
        cost_col = next((raw_cols[k] for k in raw_cols
                         if "unit inventory cost" in k or "unit_inventory_cost" in k), None)
        if cost_col and inventory_cost_df is None:
            inv_data = (raw_df[["SKU", cost_col]].drop_duplicates("SKU")
                          .rename(columns={"SKU": "Item No.", cost_col: "Unit Inventory Cost"}))
            inv_data = inv_data[inv_data["Item No."].isin(sel_skus)].reset_index(drop=True)
            # Fill any missing SKUs with 1.0
            missing = set(sel_skus) - set(inv_data["Item No."])
            if missing:
                inv_data = pd.concat([inv_data,
                    pd.DataFrame({"Item No.": list(missing), "Unit Inventory Cost": 1.0})],
                    ignore_index=True)
            inventory_cost_df = inv_data[["Item No.", "Unit Inventory Cost"]]

        # m2/item
        m2_col = next((raw_cols[k] for k in raw_cols
                       if "m2/item" in k or "m2_per_item" in k or "m2 per item" in k), None)
        if m2_col and m2_per_item_df is None:
            m2_data = (raw_df[["SKU", m2_col]].drop_duplicates("SKU")
                         .rename(columns={"SKU": "Item No.", m2_col: "m2/item"}))
            m2_data = m2_data[m2_data["Item No."].isin(sel_skus)].reset_index(drop=True)
            missing = set(sel_skus) - set(m2_data["Item No."])
            if missing:
                m2_data = pd.concat([m2_data,
                    pd.DataFrame({"Item No.": list(missing), "m2/item": 0.01})],
                    ignore_index=True)
            m2_per_item_df = m2_data[["Item No.", "m2/item"]]

    # Fallback defaults if not found in raw_df
    if inventory_cost_df is None:
        inventory_cost_df = pd.DataFrame({
            "Item No.":            sel_skus,
            "Unit Inventory Cost": [1.0] * len(sel_skus),
        })
    if m2_per_item_df is None:
        m2_per_item_df = pd.DataFrame({
            "Item No.": sel_skus,
            "m2/item":  [0.01] * len(sel_skus),
        })

    # ── Step 5: Selected_sS_Items — the filtered SKU list ────────────────────
    # Build as a 1-row-header + N-row DataFrame to reuse write_selected_sheet
    sel_df = pd.DataFrame({"Selected Item No.": ["Selected Item No."] + sel_skus})

    # ── Build data_dict — used by the optimizer ───────────────────────────────
    # The new model reads from "Predicted_Demand" and "I_ini" (matrix format).
    # data_dict keys map to what read_input_data_from_dict() expects.
    data_dict = {
        "demand_pivot":      demand_pivot,     # Predicted_Demand sheet
        "iini_pivot":        iini_pivot,       # I_ini matrix sheet (SKU × Month)
        "actual_pivot":      actual_pivot,     # Actual D_it_Imputated
        "qit_pivot":         qit_pivot,        # 2025 Q_it
        "inventory_cost":    inventory_cost_df,
        "m2_per_item":       m2_per_item_df,
        "selected_items":    sel_skus,
        "month_cols":        month_cols,       # chronological list of month numbers
        "ym_map":            ym_map,
        "ym_order":          ym_order,
    }

    # ── Step 6: Write Excel — 7 sheets matching source file ──────────────────
    wb = Workbook()

    # Sheet 1: Predicted_Demand
    ws1 = wb.active
    ws1.title = "Predicted_Demand"
    write_wide_sheet(ws1, demand_pivot, "Month", "Month")

    # Sheet 2: Actual D_it_Imputated
    ws2 = wb.create_sheet("Actual D_it_Imputated")
    write_wide_sheet(ws2, actual_pivot, "Month (year 2025)", "Month (year 2025)")

    # Sheet 3: 2025 Q_it
    ws3 = wb.create_sheet("2025 Q_it")
    write_wide_sheet(ws3, qit_pivot, "Month (year 2025)", "Month (year 2025)")

    # Sheet 4: I_ini (wide matrix)
    ws4 = wb.create_sheet("I_ini")
    write_wide_sheet(ws4, iini_pivot, "SOpening_Inventory_Qty", "SOpening_Inventory_Qty")

    # Sheet 5: Inventory Cost
    ws5 = wb.create_sheet("Inventory Cost")
    inv_df_with_header = pd.concat(
        [pd.DataFrame([inventory_cost_df.columns.tolist()], columns=inventory_cost_df.columns),
         inventory_cost_df],
        ignore_index=True,
    )
    write_narrow_sheet(ws5, inv_df_with_header)

    # Sheet 6: m2 per item
    ws6 = wb.create_sheet("m2 per item")
    m2_df_with_header = pd.concat(
        [pd.DataFrame([m2_per_item_df.columns.tolist()], columns=m2_per_item_df.columns),
         m2_per_item_df],
        ignore_index=True,
    )
    write_narrow_sheet(ws6, m2_df_with_header)

    # Sheet 7: Selected_sS_Items
    ws7 = wb.create_sheet("Selected_sS_Items")
    write_selected_sheet(ws7, sel_skus)

    buf = io.BytesIO()
    wb.save(buf)
    return data_dict, buf.getvalue()
